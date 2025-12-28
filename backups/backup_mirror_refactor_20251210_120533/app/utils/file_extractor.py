import os
from werkzeug.utils import secure_filename

class FileExtractor:
    """Extrae texto de diferentes tipos de archivos"""
    
    ALLOWED_EXTENSIONS = {'txt', 'pdf', 'xlsx', 'xls', 'docx', 'doc'}
    MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB
    
    @staticmethod
    def allowed_file(filename):
        """Verificar que la extensión es permitida"""
        return '.' in filename and \
               filename.rsplit('.', 1)[1].lower() in FileExtractor.ALLOWED_EXTENSIONS
    
    MAX_CHARS = 2000000 # Increased to 2MB text limit

    @staticmethod
    def extract_text(file_path, file_type):
        """Extraer texto según tipo de archivo"""
        
        content = ""
        
        if file_type == 'txt':
            content = FileExtractor._extract_txt(file_path)
        
        elif file_type in ['xlsx', 'xls']:
            content = FileExtractor._extract_excel(file_path)
        
        elif file_type in ['pdf']:
            content = FileExtractor._extract_pdf(file_path)
        
        elif file_type in ['docx']:
            content = FileExtractor._extract_docx(file_path)
        
        # Clean content (remove headers/footers)
        content = FileExtractor._clean_text(content)

        # Limit content length
        if len(content) > FileExtractor.MAX_CHARS:
            content = content[:FileExtractor.MAX_CHARS] + "... (truncated)"
            
        print(f"EXTRACTED TEXT ({file_type}): {len(content)} chars")
        if len(content) < 100:
            print(f"WARNING: Extracted text is very short: {content}")
            
        return content
    
    @staticmethod
    def _clean_text(text):
        """Limpiar texto de encabezados repetitivos y basura"""
        if not text: return ""
        
        # Common headers in Mexican laws
        patterns = [
            "LEY DEL IMPUESTO SOBRE LA RENTA",
            "CÁMARA DE DIPUTADOS DEL H. CONGRESO DE LA UNIÓN",
            "Secretaría General",
            "Secretaría de Servicios Parlamentarios",
            "Última Reforma DOF",
            "Texto Vigente",
            "TEXTO VIGENTE",
            "Estados Unidos Mexicanos",
            "Presidencia de la República"
        ]
        
        lines = text.split('\n')
        cleaned_lines = []
        
        for line in lines:
            line_strip = line.strip()
            # Skip empty lines
            if not line_strip:
                continue
                
            # Skip page numbers (e.g. "1 de 313")
            if " de " in line_strip and any(c.isdigit() for c in line_strip) and len(line_strip) < 20:
                continue
                
            # Skip headers
            is_header = False
            for p in patterns:
                if p in line_strip:
                    is_header = True
                    break
            
            if not is_header:
                cleaned_lines.append(line)
                
        return "\n".join(cleaned_lines)
    
    @staticmethod
    def _extract_txt(file_path):
        """Extraer texto de TXT con fallback de encoding"""
        encodings = ['utf-8', 'latin-1', 'cp1252']
        
        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    content = f.read()
                    print(f"Successfully read TXT with encoding: {encoding}")
                    return content
            except UnicodeDecodeError:
                continue
            except Exception as e:
                return f"Error leyendo archivo ({encoding}): {str(e)}"
        
        return "Error: No se pudo detectar la codificación del archivo."
    
    @staticmethod
    def _extract_excel(file_path):
        """Extraer texto de Excel"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path)
            text = []
            
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                text.append(f"\n=== HOJA: {sheet} ===\n")
                
                for row in ws.iter_rows(values_only=True):
                    text.append(" | ".join(
                        str(cell) if cell is not None else "" 
                        for cell in row
                    ))
            
            return "\n".join(text)
        
        except Exception as e:
            return f"Error leyendo Excel: {str(e)}"
    
    @staticmethod
    def _extract_pdf(file_path):
        """Extraer texto de PDF"""
        try:
            import PyPDF2
            text = []
            
            with open(file_path, 'rb') as f:
                pdf_reader = PyPDF2.PdfReader(f)
                
                for page_num, page in enumerate(pdf_reader.pages):
                    text.append(f"\n=== PÁGINA {page_num + 1} ===\n")
                    text.append(page.extract_text())
            
            return "\n".join(text)
        
        except Exception as e:
            return f"Error leyendo PDF: {str(e)}"
    
    @staticmethod
    def _extract_docx(file_path):
        """Extraer texto de DOCX"""
        try:
            from docx import Document
            doc = Document(file_path)
            text = []
            
            for para in doc.paragraphs:
                if para.text.strip():
                    text.append(para.text)
            
            return "\n".join(text)
        
        except Exception as e:
            return f"Error leyendo DOCX: {str(e)}"
