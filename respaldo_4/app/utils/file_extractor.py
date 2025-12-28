import os
from werkzeug.utils import secure_filename

class FileExtractor:
    """Extrae texto de diferentes tipos de archivos"""
    
    ALLOWED_EXTENSIONS = {'txt', 'pdf', 'xlsx', 'xls', 'docx', 'doc'}
    MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB
    
    @staticmethod
    def allowed_file(filename):
        """Verificar que la extensión es permitida"""
        return '.' in filename and \
               filename.rsplit('.', 1)[1].lower() in FileExtractor.ALLOWED_EXTENSIONS
    
    MAX_CHARS = 50000

    @staticmethod
    def extract_text(file_path, file_type):
        """Extraer texto según tipo de archivo"""
        
        content = ""
        
        if file_type == 'txt':
            content = FileExtractor._extract_txt(file_path)
        
        elif file_type in ['xlsx', 'xls']:
            content = FileExtractor._extract_excel(file_path)
        
        elif file_type in ['pdf']:
            content = FileExtractor._extract_pdf(file_path)
        
        elif file_type in ['docx']:
            content = FileExtractor._extract_docx(file_path)
        
        # Limit content length
        if len(content) > FileExtractor.MAX_CHARS:
            content = content[:FileExtractor.MAX_CHARS] + "... (truncated)"
            
        return content
    
    @staticmethod
    def _extract_txt(file_path):
        """Extraer texto de TXT"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
        except Exception as e:
            return f"Error leyendo archivo: {str(e)}"
    
    @staticmethod
    def _extract_excel(file_path):
        """Extraer texto de Excel"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path)
            text = []
            
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                text.append(f"\n=== HOJA: {sheet} ===\n")
                
                for row in ws.iter_rows(values_only=True):
                    text.append(" | ".join(
                        str(cell) if cell is not None else "" 
                        for cell in row
                    ))
            
            return "\n".join(text)
        
        except Exception as e:
            return f"Error leyendo Excel: {str(e)}"
    
    @staticmethod
    def _extract_pdf(file_path):
        """Extraer texto de PDF"""
        try:
            import PyPDF2
            text = []
            
            with open(file_path, 'rb') as f:
                pdf_reader = PyPDF2.PdfReader(f)
                
                for page_num, page in enumerate(pdf_reader.pages):
                    text.append(f"\n=== PÁGINA {page_num + 1} ===\n")
                    text.append(page.extract_text())
            
            return "\n".join(text)
        
        except Exception as e:
            return f"Error leyendo PDF: {str(e)}"
    
    @staticmethod
    def _extract_docx(file_path):
        """Extraer texto de DOCX"""
        try:
            from docx import Document
            doc = Document(file_path)
            text = []
            
            for para in doc.paragraphs:
                if para.text.strip():
                    text.append(para.text)
            
            return "\n".join(text)
        
        except Exception as e:
            return f"Error leyendo DOCX: {str(e)}"
